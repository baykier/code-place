# coding:utf-8

from openpyxl import load_workbook
import os
from datetime import date
from datetime import timedelta

"""
根据给的的参数，自动生成周报
"""

# 默认配置

MODEL_FILE = os.path.dirname(os.path.abspath(__file__)) + '/week_report_model.xlsx'
REPORT_NAME = '张志杰'
JOB = [
    ('crm优化ajax没权限提示信息', 1, 1),  # tuple 第一个元素 干的啥 第二个 开始时间 第三个 结束时间
    ('划外绩效申请新功能规划', 2, 2),
    ('划外绩效申请申请表单完成', 3, 3),
    ('划外绩效申请申请表单列表页完成', 4, 5)
]

print(MODEL_FILE)

def get_week_date(di):
    '''
    获取本周一周之内的日期
    di 一周之内的数字索引 1 - 7 周一对应1 周五对应5 周日对应7
    '''
    return date.today() - timedelta(date.today().isoweekday() - di)

def get_report_save_file(name = REPORT_NAME):
    '''
    获取周报的文件名
    '''
    friday_date =  get_week_date(5).strftime('%Y%m%d')
    return '技术工作周报-网利宝技术团队{0}{1}.xlsx'.format(name, friday_date)
    
def parse_jobs(jobs=JOB):
    '''
    解析JOB列表 返回要写入的报告列表
    '''

    _jobs = []
    for index, _job in enumerate(jobs, 6):     
        _tuple = [
            ('E' + str(index), _job[0], index),
            ('F' + str(index), get_week_date(_job[1]).strftime("%-m/%d"), index),
            ('H' + str(index), get_week_date(_job[2]).strftime("%-m/%d"), index)
        ]
        _jobs.append(_tuple)
    return _jobs


def auto_report(name = REPORT_NAME, model=MODEL_FILE):
    '''
    根据给定的模板和参数，自动生成周报
    '''
    pass

    
    if not JOB:
        print('工作任务列表为空')
    else:
        wb = load_workbook(filename=model)
        ws = wb.active

        ws['I4'] = get_week_date(5).strftime('%Y-%m-%d')  # 提交日期
        ws['G4'] = name #报告人      
        jobs = parse_jobs()
        for job in jobs:
            for k,v,i in job:
                ws[str(k)] = str(v)
                ws['I' + str(i)] = '完成'
        wb.save(get_report_save_file())
        print('周报自动生成ok!')

def main():
    auto_report()

if __name__ == '__main__':   
    main()

